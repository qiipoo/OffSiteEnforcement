# -*- coding: utf-8 -*-

import os
import pathlib
import re
from openpyxl import Workbook, worksheet, cell, load_workbook


def open_file():

    work_path = 'C:\\Users\\JT-0919\\Desktop\\WORK\\执法项目\\数据'

    os.chdir(work_path)

    wb = Workbook()
    wb = load_workbook('数据中心-信息资源表设计-20220829.xlsx')

    pattern_table_name = r'\w+'
    pattern_table_desc = r'\W[^（）]+'

    for sheet in wb.worksheets:
        if sheet.title == '目录':
            continue

        table_name = sheet['A1'].value
        ret = re.findall(pattern_table_name, table_name)

        table_name = ret[1]
        table_desc = ret[0]

        fields = sheet['2']

        columns = []
        for one_row in sheet.iter_rows(min_row=3):
            col_nm = ''
            col_tp = ''
            col_nl = ''
            col_cmt = ''
            for one_cell in one_row:
                for field in fields:
                    if field.column == one_cell.column:
                        if field.value == '字段':
                            col_nm = one_cell.value
                        elif field.value == '字段类型' or field.value == '类型' \
                                or field.value == '数据类型' or field.value == '数据类型及长度':
                            col_tp = one_cell.value
                            if type(col_tp) == str:
                                col_tp = col_tp.replace('VARCHAR2', 'VARCHAR')
                                col_tp = col_tp.replace('（', '(')
                                col_tp = col_tp.replace('）', ')')
                                col_tp = col_tp.replace('STRING', 'VARCHAR')
                                col_tp = col_tp.replace('VARCHAR2', 'VARCHAR')
                                col_tp = col_tp.replace('VARCHAR2', 'VARCHAR')
                        elif field.value == '长度':
                            col_tp = '{}({})'.format(col_tp, one_cell.value)
                        elif field.value == '字段注释' or field.value == '信息名称':
                            col_cmt = one_cell.value
                        elif field.value == '备注' or field.value == '是否允许为空':
                            if one_cell.value == '必填' or one_cell.value == '否':
                                col_nl = 'NOT NULL'

            column = ColumnDef(col_nm, col_tp, col_nl, col_cmt)
            columns.append(column)

        one_table = TableDef(table_name, columns, table_desc)

        filedir = os.path.join(work_path, 'DDL')
        if not os.path.exists(filedir):
            os.mkdir(filedir)
        filepath = os.path.join(filedir, sheet.title + '.txt')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("CREATE TABLE %s (\n" % one_table.table_name)
            cnt = 0
            for one_col in one_table.table_columns:
                cnt = cnt + 1
                f.write(" %-22s" % one_col.col_name)
                f.write(" %-13s" % one_col.col_type)
                f.write(" %-8s" % one_col.col_not_null)
                if cnt == len(one_table.table_columns):
                    f.write(" COMMENT '%s'\n" % one_col.col_comment)
                else:
                    f.write(" COMMENT '%s',\n" % one_col.col_comment)
            f.write(") \n COMMENT='%s';" % one_table.table_comment)


class ColumnDef:
    col_index = ''
    col_name = ''
    col_type = ''
    col_len = ''
    col_not_null = ''
    col_default_val = ''
    col_key = ''
    col_comment = ''

    def __init__(self, name, typ, null, comment):
        self.col_name = name
        self.col_type = typ
        self.col_not_null = null
        self.col_comment = comment


class TableDef(ColumnDef):
    table_name = ''
    table_columns = []
    table_comment = ''

    def __init__(self, name, columns, comment):
        self.table_name = name
        self.table_columns = columns
        self.table_comment = comment
