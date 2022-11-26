# -*- coding: utf-8 -*-
import os.path
import pandas as pd
import numpy as np
import math
import copy
import time
import datetime
import re
import glob
import openpyxl

# 文件定义
input_path = r'C:\Users\JT-0919\Desktop\WORK\执法项目\09.数据\20221122'
output_path = r'C:\Users\JT-0919\Desktop\WORK\执法项目\09.数据\OUTPUT'
except_path = r'C:\Users\JT-0919\Desktop\WORK\执法项目\09.数据\20221115\新建文件夹'

input_mj_path = r'C:\Users\JT-0919\Desktop\WORK\执法项目\09.数据\20221122\result'

input_file = r'20221010-20221110(原始).csv'
# input_file = r'20221010-20221110(剔除包车).csv'

output_total_file = r'车辆汇总表.csv'
output_detail_file_csv = r'车辆明细表.csv'
output_detail_file_xlsx = r'车辆明细表.xlsx'

except_file1 = r'2.黑客运.csv'
except_file2 = r'2.黑旅游.csv'
except_file3 = r'3.11月3日-10日环京疫情重点地区进京客运车辆监测统计表-发送版.csv'
except_file4 = r'3.11月12日环京疫情重点地区进京客运车辆监测统计表-发送版.csv'
except_file5 = r'3.11月15日客运进京日表-发送版.csv'
except_file6 = r'3.业内客车基础信息.csv'
except_file7 = r'4.公交车辆信息.csv'


#  'sf',  'ch',  'cx',   'rk',   'rksj',    'ck',    'gs',     'cksj'
# '省份', '车牌', '车型', '入口站', '入口时间', '出口站', '所在高速', '出口时间'
input_columns = ['sf', 'ch', 'cx', 'rk', 'rksj', 'ck', 'gs', 'cksj']
input_data = pd.DataFrame(data=None, columns=input_columns)

input_files = glob.glob(os.path.join(input_path, input_file))
input_mj_files = glob.glob(os.path.join(input_mj_path, '*'))

# 文件读取
input_data_lst = []
for idx, in_file in enumerate(input_files):
    in_data = pd.read_csv(in_file, encoding="utf-8", keep_default_na=False)
    input_data_lst.append(in_data)
input_data = pd.concat(input_data_lst)
except_data1 = pd.read_csv(os.path.join(except_path, except_file1), encoding="utf-8", na_filter=False)
except_data2 = pd.read_csv(os.path.join(except_path, except_file2), encoding="utf-8", na_filter=False)
except_data3 = pd.read_csv(os.path.join(except_path, except_file3), encoding="utf-8", na_filter=False)
except_data4 = pd.read_csv(os.path.join(except_path, except_file4), encoding="utf-8", na_filter=False)
except_data5 = pd.read_csv(os.path.join(except_path, except_file5), encoding="utf-8", na_filter=False)
except_data6 = pd.read_csv(os.path.join(except_path, except_file6), encoding="utf-8", na_filter=False)
except_data7 = pd.read_csv(os.path.join(except_path, except_file7), encoding="utf-8", na_filter=False)

mj_columns = ['通行标识ID', '车牌号', '车牌类型', '计费车型代码', '通过时间', '门架编号', '匹配门架名称', '匹配门架经度', '匹配门架纬度']
input_mj_data = pd.DataFrame(data=None, columns=mj_columns)
input_mj_data_lst = []
for idx, in_file in enumerate(input_mj_files):
    in_data = pd.read_csv(in_file, encoding="utf-8", keep_default_na=False, names=mj_columns)
    input_mj_data_lst.append(in_data)
input_mj_data = pd.concat(input_mj_data_lst)

# 过滤条件
set_exp_data1 = set(except_data1['车辆牌号'].tolist())
set_exp_data2 = set(except_data2['车辆牌号'].tolist())
set_exp_data3 = set(except_data3['车牌号'].tolist())
set_exp_data4 = set(except_data4['车牌号'].tolist())
set_exp_data5 = set(except_data5['车牌号'].tolist())
set_exp_data6 = set(except_data6['车牌号码'].tolist())
set_exp_data7 = set(except_data7['牌照号'].tolist())

# '客一', '客二', '客三', '客四'
lst_cx = ['客一']
lst_rk = []
lst_ck = []
set_cx = set(lst_cx)
set_rk = set(lst_rk)
set_ck = set(lst_ck)
filter_dt = datetime.datetime(2022, 10, 10)
filter_cnt = 15

# 统计操作
output_total_data = []
output_detail_data = []
output_filter_detail_data = []
output_mj_detail_data = []
output_data_cnt = {}
for idx, itr in enumerate(input_data.itertuples(), start=1):
    # print("[%s] Loading [%d/%d]" % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), (idx + 1), len(input_data)))

    sf = itr[1]
    ch = itr[2]
    cx = itr[3]
    rk = itr[4]
    rksj = itr[5]
    ck = itr[6]
    gs = itr[7]
    cksj = itr[8]

    rksj_dt = ''
    rksj_tm = ''
    rksj_hour = ''
    cksj_dt = ''
    cksj_tm = ''
    cksj_hour = ''

    yn = ''
    yz = ''
    hky_txt = ''
    hky_dt = ''
    hly_txt = ''
    hly_dt = ''

    if cx in set_cx:
        continue

    if rk in set_rk:
        continue

    if ck in set_ck:
        continue

    if isinstance(ch, float) or ch == '':
        print('Skip [%s, %s, %s, %s, %s, %s, %s, %s, %s, %s]' % itr)
        continue

    if isinstance(cx, float) or cx == '':
        print('Skip [%s, %s, %s, %s, %s, %s, %s, %s, %s, %s]' % itr)
        continue

    if isinstance(rksj, str):
        rksj_dt = str.split(rksj, ' ')[0]
        if len(str.split(rksj, ' ')) == 2:
            rksj_tm = str.split(rksj, ' ')[1]
            rksj_tm = str.split(rksj_tm, '.')[0]
            rksj_hour = str.split(rksj_tm, ':')[0]

    if isinstance(cksj, str):
        cksj_dt = str.split(cksj, ' ')[0]
        if len(str.split(cksj, ' ')) == 2:
            cksj_tm = str.split(cksj, ' ')[1]
            cksj_tm = str.split(cksj_tm, '.')[0]
            cksj_hour = str.split(cksj_tm, ':')[0]

    if datetime.datetime.strptime(cksj_dt, '%Y-%m-%d') < filter_dt:
        print(itr)
        continue

    if ch in set_exp_data1:
        hky_txt = except_data1.loc[except_data1['车辆牌号'] == ch].values[0][6]
        hky_dt = except_data1.loc[except_data1['车辆牌号'] == ch].values[0][7]

    if ch in set_exp_data2:
        hly_txt = except_data2.loc[except_data2['车辆牌号'] == ch].values[0][7]
        hly_dt = except_data2.loc[except_data2['车辆牌号'] == ch].values[0][8]

    if ch in set_exp_data3:
        yn = ''
        yz = except_data3.loc[except_data3['车牌号'] == ch].values[0][11]
        if re.search(r'租赁', yz) or len(yz) == 2 or len(yz) == 3:
            yn = '疑似'
        else:
            yn = ''

    if ch in set_exp_data4:
        yn = ''
        yz = except_data4.loc[except_data4['车牌号'] == ch].values[0][11]
        if re.search(r'租赁', yz) or len(yz) == 2 or len(yz) == 3:
            yn = '疑似'
        else:
            yn = ''

    if ch in set_exp_data5:
        yn = ''
        yz = except_data5.loc[except_data5['车牌号'] == ch].values[0][11]
        if re.search(r'租赁', yz) or len(yz) == 2 or len(yz) == 3:
            yn = '疑似'
        else:
            yn = ''

    if ch in set_exp_data6:
        yn = ''
        yz = except_data6.loc[except_data6['车牌号码'] == ch].values[0][7]
        if re.search(r'租赁', yz) or len(yz) == 2 or len(yz) == 3:
            yn = '疑似'
        else:
            yn = '业内'

    if ch in set_exp_data7:
        yn = '业内'
        yz = '北京公交集团'

    cnt_value = output_data_cnt.get(ch)
    if cnt_value is None:
        cnt_inf = {ch: 1}
    else:
        cnt_inf = {ch: cnt_value + 1}
    output_data_cnt.update(cnt_inf)

    output_detail_data.append([sf, ch, cx, rk, rksj_dt, rksj_tm, rksj_hour, ck, gs, cksj_dt, cksj_tm, cksj_hour,
                               yn, yz, hky_txt, hky_dt, hly_txt, hly_dt])



# 过滤
filter_data_cnt = dict(filter(lambda x: x[1] >= filter_cnt, output_data_cnt.items()))

# 排序
sorted_data_cnt = sorted(filter_data_cnt.items(), key=lambda x: x[1], reverse=True)

# 过滤后明细
for a_detail_data in output_detail_data:
    if a_detail_data[1] in filter_data_cnt.keys():
        output_filter_detail_data.append(a_detail_data)

for a_sorted_data in sorted_data_cnt:
    for a_filter_detail_data in output_filter_detail_data:
        if a_sorted_data[0] == a_filter_detail_data[1]:
            a_output_total_data = {
                '车牌': a_sorted_data[0],
                '次数': a_sorted_data[1],
                '疑似业内': a_filter_detail_data[12],
                '所有人': a_filter_detail_data[13],
                '黑客运违法情形': a_filter_detail_data[14],
                '黑客运检查时间': a_filter_detail_data[15],
                '黑旅游违法情形': a_filter_detail_data[16],
                '黑旅游检查时间': a_filter_detail_data[17]
            }
    output_total_data.append(a_output_total_data)


for a_mj_data in input_mj_data.itertuples():
    if a_mj_data[2] in filter_data_cnt.keys():
        output_mj_detail_data.append(a_mj_data)

# csv明细文件输出
csv_detail_title = ('省份', '车牌', '车型',
                    '入口', '入口日期', '入口时间', '入口几点',
                    '出口', '所属高速', '出口日期', '出口时间', '出口几点',
                    '疑似业内', '所有人', '黑客运违法情形', '黑客运检查时间', '黑旅游违法情形', '黑旅游检查时间')

with open(os.path.join(output_path, output_detail_file_csv), 'w') as f:
    f.write('%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s\n' % csv_detail_title)

    for idx, a_detail_data in enumerate(output_detail_data):
        f.write(','.join(a_detail_data))
        f.write('\n')


# excel明细文件输出
data_columns = ['省份', '车牌', '车型',
                '入口', '入口日期', '入口时间', '入口几点',
                '出口', '所属高速', '出口日期', '出口时间', '出口几点',
                '疑似业内', '所有人', '黑客运违法情形', '黑客运检查时间', '黑旅游违法情形', '黑旅游检查时间']
excel_file_path = os.path.join(output_path, output_detail_file_xlsx)
# df_out_value = pd.DataFrame.from_dict(output_detail_data, orient='index', columns=data_columns)
df_out_value = pd.DataFrame.from_dict(output_detail_data)
with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    df_out_value.to_excel(writer, sheet_name='全体车辆明细', index=False)

df_out_value = pd.DataFrame.from_dict(output_filter_detail_data)
with pd.ExcelWriter(excel_file_path, mode='a', engine='openpyxl') as writer:
    df_out_value.to_excel(writer, sheet_name='15次以上车辆明细', index=False)

df_out_value = pd.DataFrame.from_dict(output_total_data)
with pd.ExcelWriter(excel_file_path, mode='a', engine='openpyxl') as writer:
    df_out_value.to_excel(writer, sheet_name='15次以上车辆统计', index=False)

df_out_value = pd.DataFrame.from_dict(output_mj_detail_data)
with pd.ExcelWriter(excel_file_path, mode='a', engine='openpyxl') as writer:
    df_out_value.to_excel(writer, sheet_name='15次以上车辆门架明细', index=False)

wb = openpyxl.load_workbook(excel_file_path)
sheet = wb['全体车辆明细']
sheet.cell(column=1, row=1, value='省份')
sheet.cell(column=2, row=1, value='车牌')
sheet.cell(column=3, row=1, value='车型')
sheet.cell(column=4, row=1, value='入口')
sheet.cell(column=5, row=1, value='入口日期')
sheet.cell(column=6, row=1, value='入口时间')
sheet.cell(column=7, row=1, value='入口几点')
sheet.cell(column=8, row=1, value='出口')
sheet.cell(column=9, row=1, value='所属高速')
sheet.cell(column=10, row=1, value='出口日期')
sheet.cell(column=11, row=1, value='出口时间')
sheet.cell(column=12, row=1, value='出口几点')
sheet.cell(column=13, row=1, value='疑似业内')
sheet.cell(column=14, row=1, value='所有人')
sheet.cell(column=15, row=1, value='黑客运违法情形')
sheet.cell(column=16, row=1, value='黑客运检查时间')
sheet.cell(column=17, row=1, value='黑旅游违法情形')
sheet.cell(column=18, row=1, value='黑旅游检查时间')

sheet = wb['15次以上车辆明细']
sheet.cell(column=1, row=1, value='省份')
sheet.cell(column=2, row=1, value='车牌')
sheet.cell(column=3, row=1, value='车型')
sheet.cell(column=4, row=1, value='入口')
sheet.cell(column=5, row=1, value='入口日期')
sheet.cell(column=6, row=1, value='入口时间')
sheet.cell(column=7, row=1, value='入口几点')
sheet.cell(column=8, row=1, value='出口')
sheet.cell(column=9, row=1, value='所属高速')
sheet.cell(column=10, row=1, value='出口日期')
sheet.cell(column=11, row=1, value='出口时间')
sheet.cell(column=12, row=1, value='出口几点')
sheet.cell(column=13, row=1, value='疑似业内')
sheet.cell(column=14, row=1, value='所有人')
sheet.cell(column=15, row=1, value='黑客运违法情形')
sheet.cell(column=16, row=1, value='黑客运检查时间')
sheet.cell(column=17, row=1, value='黑旅游违法情形')
sheet.cell(column=18, row=1, value='黑旅游检查时间')

wb.save(excel_file_path)
